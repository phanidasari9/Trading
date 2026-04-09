import io
import warnings
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)

# ============================================================
# CONFIG
# ============================================================
TICKERS = [
    "AAPL", "MSFT", "NVDA", "AMZN", "META", "TSLA", "AMD", "NFLX",
    "GOOGL", "PLTR", "AVGO", "MU", "SMCI", "SPY", "QQQ"
]

DAILY_PERIOD = "3mo"
INTRADAY_PERIOD = "5d"
INTRADAY_INTERVAL = "5m"

MIN_PRICE = 5
MIN_AVG_DOLLAR_VOL = 20_000_000
MIN_OPTION_VOLUME = 100
MIN_OPEN_INTEREST = 100
UNUSUAL_VOL_OI = 2.0
TOP_N = 10

OUTPUT_FILE = f"money_flow_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


# ============================================================
# HELPERS
# ============================================================
def safe_float(x, default=np.nan):
    try:
        if x is None:
            return default
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
    return df


def compute_obv(close: pd.Series, volume: pd.Series) -> pd.Series:
    direction = np.sign(close.diff().fillna(0))
    direction = direction.replace(0, np.nan).ffill().fillna(0)
    return (direction * volume).cumsum()


def compute_cmf(df: pd.DataFrame, period: int = 20) -> pd.Series:
    hl_range = (df["High"] - df["Low"]).replace(0, np.nan)
    mf_multiplier = ((df["Close"] - df["Low"]) - (df["High"] - df["Close"])) / hl_range
    mf_volume = mf_multiplier.fillna(0) * df["Volume"]
    return mf_volume.rolling(period).sum() / df["Volume"].rolling(period).sum()


def add_intraday_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Bar_Return_%"] = df["Close"].pct_change() * 100
    df["Dollar_Volume"] = df["Close"] * df["Volume"]
    df["Vol_MA_20"] = df["Volume"].rolling(20).mean()
    df["Rel_Vol"] = df["Volume"] / df["Vol_MA_20"]
    df["Flow_Score"] = df["Bar_Return_%"].fillna(0) * df["Rel_Vol"].fillna(0)
    return df


def add_daily_features(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Pct_Change_%"] = df["Close"].pct_change() * 100
    df["Dollar_Volume"] = df["Close"] * df["Volume"]
    df["Avg_Vol_20"] = df["Volume"].rolling(20).mean()
    df["Avg_Dollar_Vol_20"] = df["Dollar_Volume"].rolling(20).mean()
    df["Rel_Vol"] = df["Volume"] / df["Avg_Vol_20"]
    df["CMF_20"] = compute_cmf(df, 20)
    df["OBV"] = compute_obv(df["Close"], df["Volume"])
    df["OBV_Slope_5"] = df["OBV"].diff(5)
    df["Money_Flow_Score"] = df["Pct_Change_%"].fillna(0) * df["Rel_Vol"].fillna(0)
    return df


def get_last_valid(series: pd.Series, default=np.nan):
    s = series.dropna()
    if s.empty:
        return default
    return s.iloc[-1]


# ============================================================
# CORE ANALYSIS
# ============================================================
def fetch_daily_data(ticker: str) -> pd.DataFrame:
    df = yf.download(
        ticker,
        period=DAILY_PERIOD,
        interval="1d",
        auto_adjust=True,
        progress=False,
        threads=False
    )
    if df is None or df.empty:
        return pd.DataFrame()
    return add_daily_features(flatten_columns(df))


def fetch_intraday_data(ticker: str) -> pd.DataFrame:
    df = yf.download(
        ticker,
        period=INTRADAY_PERIOD,
        interval=INTRADAY_INTERVAL,
        auto_adjust=True,
        progress=False,
        threads=False
    )
    if df is None or df.empty:
        return pd.DataFrame()
    return add_intraday_features(flatten_columns(df))


def summarize_intraday_flow(intraday_df: pd.DataFrame) -> dict:
    if intraday_df.empty or len(intraday_df) < 10:
        return {
            "Intraday_Bull_Flow": np.nan,
            "Intraday_Bear_Flow": np.nan,
            "Intraday_Net_Flow": np.nan,
            "Last_5m_Return_%": np.nan,
            "Last_5m_RelVol": np.nan
        }

    bull_flow = intraday_df.loc[intraday_df["Flow_Score"] > 0, "Flow_Score"].sum()
    bear_flow = intraday_df.loc[intraday_df["Flow_Score"] < 0, "Flow_Score"].sum()
    net_flow = intraday_df["Flow_Score"].sum()

    latest = intraday_df.iloc[-1]
    return {
        "Intraday_Bull_Flow": round(float(bull_flow), 2),
        "Intraday_Bear_Flow": round(float(bear_flow), 2),
        "Intraday_Net_Flow": round(float(net_flow), 2),
        "Last_5m_Return_%": round(safe_float(latest.get("Bar_Return_%")), 3),
        "Last_5m_RelVol": round(safe_float(latest.get("Rel_Vol")), 2)
    }


def get_option_side_summary(chain_df: pd.DataFrame, side: str, spot: float) -> pd.DataFrame:
    if chain_df is None or chain_df.empty:
        return pd.DataFrame()

    df = chain_df.copy()

    for col in ["volume", "openInterest", "bid", "ask", "lastPrice", "strike", "impliedVolatility"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["mid"] = np.where(
        (df["bid"].fillna(0) > 0) & (df["ask"].fillna(0) > 0),
        (df["bid"].fillna(0) + df["ask"].fillna(0)) / 2,
        df["lastPrice"]
    )

    df["dollar_premium"] = df["mid"].fillna(0) * df["volume"].fillna(0) * 100
    df["vol_oi_ratio"] = df["volume"] / df["openInterest"].replace(0, np.nan)
    df["distance_from_spot_%"] = ((df["strike"] - spot) / spot) * 100

    # Approx nearest ATM clustering
    df["abs_distance"] = (df["strike"] - spot).abs()

    # "Bullish" side = calls, "Bearish" side = puts by convention here
    if side.lower() == "calls":
        df["flow_bias"] = "Bullish"
    else:
        df["flow_bias"] = "Bearish"

    # Unusual filters
    filtered = df[
        (df["volume"].fillna(0) >= MIN_OPTION_VOLUME) &
        (df["openInterest"].fillna(0) >= MIN_OPEN_INTEREST)
    ].copy()

    if filtered.empty:
        return pd.DataFrame()

    filtered["is_unusual"] = (
        (filtered["volume"].fillna(0) >= 1000) |
        (filtered["vol_oi_ratio"].fillna(0) >= UNUSUAL_VOL_OI)
    )

    filtered = filtered.sort_values(
        by=["is_unusual", "dollar_premium", "volume"],
        ascending=[False, False, False]
    )

    keep_cols = [
        "contractSymbol", "strike", "lastPrice", "bid", "ask", "mid",
        "volume", "openInterest", "vol_oi_ratio", "impliedVolatility",
        "dollar_premium", "distance_from_spot_%", "flow_bias", "is_unusual"
    ]
    keep_cols = [c for c in keep_cols if c in filtered.columns]
    return filtered[keep_cols].head(TOP_N).reset_index(drop=True)


def analyze_options_flow(ticker: str, spot: float) -> tuple[dict, pd.DataFrame]:
    tk = yf.Ticker(ticker)
    expiries = getattr(tk, "options", []) or []

    if not expiries:
        return (
            {
                "Nearest_Expiry": None,
                "Call_Premium_Flow": np.nan,
                "Put_Premium_Flow": np.nan,
                "Net_Option_Flow": np.nan,
                "Call_Volume": np.nan,
                "Put_Volume": np.nan,
                "Put_Call_Vol_Ratio": np.nan,
                "Top_Options_Bias": None
            },
            pd.DataFrame()
        )

    # prioritize nearest expiry
    expiry = expiries[0]

    try:
        opt = tk.option_chain(expiry)
        calls = opt.calls.copy()
        puts = opt.puts.copy()
    except Exception:
        return (
            {
                "Nearest_Expiry": expiry,
                "Call_Premium_Flow": np.nan,
                "Put_Premium_Flow": np.nan,
                "Net_Option_Flow": np.nan,
                "Call_Volume": np.nan,
                "Put_Volume": np.nan,
                "Put_Call_Vol_Ratio": np.nan,
                "Top_Options_Bias": None
            },
            pd.DataFrame()
        )

    calls_top = get_option_side_summary(calls, "calls", spot)
    puts_top = get_option_side_summary(puts, "puts", spot)

    def premium_sum(df: pd.DataFrame) -> float:
        if df.empty or "dollar_premium" not in df.columns:
            return 0.0
        return float(df["dollar_premium"].sum())

    call_premium = premium_sum(calls_top)
    put_premium = premium_sum(puts_top)

    call_volume = float(calls["volume"].fillna(0).sum()) if not calls.empty else 0.0
    put_volume = float(puts["volume"].fillna(0).sum()) if not puts.empty else 0.0
    pcr = (put_volume / call_volume) if call_volume > 0 else np.nan

    combined = pd.concat(
        [
            calls_top.assign(OptionType="CALL") if not calls_top.empty else pd.DataFrame(),
            puts_top.assign(OptionType="PUT") if not puts_top.empty else pd.DataFrame()
        ],
        ignore_index=True
    )

    if not combined.empty and "dollar_premium" in combined.columns:
        combined = combined.sort_values(by=["is_unusual", "dollar_premium"], ascending=[False, False]).reset_index(drop=True)

    if call_premium > put_premium:
        top_bias = "Bullish"
    elif put_premium > call_premium:
        top_bias = "Bearish"
    else:
        top_bias = "Neutral"

    summary = {
        "Nearest_Expiry": expiry,
        "Call_Premium_Flow": round(call_premium, 2),
        "Put_Premium_Flow": round(put_premium, 2),
        "Net_Option_Flow": round(call_premium - put_premium, 2),
        "Call_Volume": int(call_volume),
        "Put_Volume": int(put_volume),
        "Put_Call_Vol_Ratio": round(pcr, 2) if not pd.isna(pcr) else np.nan,
        "Top_Options_Bias": top_bias
    }

    return summary, combined.head(20)


def analyze_ticker(ticker: str) -> tuple[dict | None, pd.DataFrame]:
    try:
        daily = fetch_daily_data(ticker)
        if daily.empty or len(daily) < 25:
            return None, pd.DataFrame()

        latest = daily.iloc[-1]
        close = safe_float(latest["Close"])
        avg_dollar_vol = safe_float(latest["Avg_Dollar_Vol_20"])

        if pd.isna(close) or close < MIN_PRICE:
            return None, pd.DataFrame()
        if pd.isna(avg_dollar_vol) or avg_dollar_vol < MIN_AVG_DOLLAR_VOL:
            return None, pd.DataFrame()

        intraday = fetch_intraday_data(ticker)
        intraday_summary = summarize_intraday_flow(intraday)
        options_summary, options_detail = analyze_options_flow(ticker, close)

        total_score = (
            safe_float(latest["Money_Flow_Score"], 0)
            + 0.5 * safe_float(intraday_summary["Intraday_Net_Flow"], 0)
            + (safe_float(options_summary["Net_Option_Flow"], 0) / 1_000_000)
        )

        direction = "Bullish Inflow" if total_score > 0 else "Bearish Outflow"

        summary = {
            "Ticker": ticker,
            "Close": round(close, 2),
            "Daily_Change_%": round(safe_float(latest["Pct_Change_%"]), 2),
            "Volume": int(safe_float(latest["Volume"], 0)),
            "Rel_Vol": round(safe_float(latest["Rel_Vol"]), 2),
            "Dollar_Volume": round(safe_float(latest["Dollar_Volume"]), 2),
            "Avg_Dollar_Vol_20": round(avg_dollar_vol, 2),
            "CMF_20": round(safe_float(latest["CMF_20"]), 3),
            "OBV_Slope_5": round(safe_float(latest["OBV_Slope_5"]), 2),
            "Daily_Money_Flow_Score": round(safe_float(latest["Money_Flow_Score"]), 2),
            **intraday_summary,
            **options_summary,
            "Composite_Flow_Score": round(total_score, 2),
            "Direction": direction
        }
        return summary, options_detail

    except Exception as e:
        print(f"[ERROR] {ticker}: {e}")
        return None, pd.DataFrame()


# ============================================================
# EXCEL OUTPUT
# ============================================================
def autosize_worksheet(ws):
    for col in ws.columns:
        max_len = 0
        col_idx = col[0].column
        for cell in col:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 24)


def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    autosize_worksheet(ws)


def apply_conditional_formatting(ws):
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    if "Composite_Flow_Score" in headers:
        col = get_column_letter(headers["Composite_Flow_Score"])
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )
        )

    for field in ["Daily_Money_Flow_Score", "Intraday_Net_Flow", "Net_Option_Flow"]:
        if field in headers:
            col = get_column_letter(headers[field])
            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="num", mid_value=0, mid_color="FFF2CC",
                    end_type="max", end_color="63BE7B"
                )
            )

    if "Direction" in headers:
        col = get_column_letter(headers["Direction"])
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(operator="equal", formula=['"Bullish Inflow"'],
                       fill=PatternFill("solid", fgColor="C6EFCE"))
        )
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(operator="equal", formula=['"Bearish Outflow"'],
                       fill=PatternFill("solid", fgColor="FFC7CE"))
        )


def _write_flow_sheets(
    writer: pd.ExcelWriter,
    master_df: pd.DataFrame,
    bullish_df: pd.DataFrame,
    bearish_df: pd.DataFrame,
    options_map: dict,
) -> None:
    master_df.to_excel(writer, sheet_name="All_Flows", index=False)
    bullish_df.to_excel(writer, sheet_name="Bullish_Inflow", index=False)
    bearish_df.to_excel(writer, sheet_name="Bearish_Outflow", index=False)

    used_names: set[str] = {"All_Flows", "Bullish_Inflow", "Bearish_Outflow"}
    for ticker, opt_df in options_map.items():
        if opt_df is not None and not opt_df.empty:
            base = f"{ticker}_Options"[:31]
            name = base
            n = 0
            while name in used_names:
                n += 1
                suffix = f"_{n}"
                name = (f"{ticker}_Options")[: 31 - len(suffix)] + suffix
            used_names.add(name)
            opt_df.to_excel(writer, sheet_name=name, index=False)


def _style_flow_workbook(wb) -> None:
    for ws in wb.worksheets:
        style_sheet(ws)
        if ws.title in {"All_Flows", "Bullish_Inflow", "Bearish_Outflow"}:
            apply_conditional_formatting(ws)


def write_excel(
    master_df: pd.DataFrame,
    bullish_df: pd.DataFrame,
    bearish_df: pd.DataFrame,
    options_map: dict,
    output_path: str | Path,
) -> None:
    output_path = Path(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_flow_sheets(writer, master_df, bullish_df, bearish_df, options_map)

    wb = load_workbook(output_path)
    _style_flow_workbook(wb)
    wb.save(output_path)


def money_flow_excel_bytes(
    master_df: pd.DataFrame,
    bullish_df: pd.DataFrame,
    bearish_df: pd.DataFrame,
    options_map: dict,
) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_flow_sheets(writer, master_df, bullish_df, bearish_df, options_map)
    buf.seek(0)
    wb = load_workbook(buf)
    _style_flow_workbook(wb)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def run_money_flow_scan(
    tickers: list[str],
    top_n: int | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, pd.DataFrame]]:
    """
    Daily + intraday + nearest-expiry options money flow per symbol.
    Returns (all rows sorted by Composite_Flow_Score, top bullish, top bearish, options detail per ticker).
    """
    tn = TOP_N if top_n is None else top_n
    summaries: list[dict] = []
    options_map: dict[str, pd.DataFrame] = {}

    for ticker in tickers:
        summary, options_detail = analyze_ticker(ticker)
        if summary:
            summaries.append(summary)
            options_map[ticker] = options_detail

    if not summaries:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), {}

    df = pd.DataFrame(summaries)
    df = df.sort_values(
        by=["Composite_Flow_Score", "Dollar_Volume"],
        ascending=[False, False],
    ).reset_index(drop=True)

    bullish_df = (
        df[df["Composite_Flow_Score"] > 0]
        .sort_values(by="Composite_Flow_Score", ascending=False)
        .head(tn)
        .reset_index(drop=True)
    )
    bearish_df = (
        df[df["Composite_Flow_Score"] < 0]
        .sort_values(by="Composite_Flow_Score", ascending=True)
        .head(tn)
        .reset_index(drop=True)
    )

    return df, bullish_df, bearish_df, options_map


# ============================================================
# MAIN
# ============================================================
def main():
    print("Scanning tickers...")
    df, bullish_df, bearish_df, options_map = run_money_flow_scan(TICKERS, top_n=TOP_N)

    if df.empty:
        print("No valid results found.")
        return

    write_excel(df, bullish_df, bearish_df, options_map, OUTPUT_FILE)

    print("\n=== TOP BULLISH INFLOW ===")
    if bullish_df.empty:
        print("None")
    else:
        print(bullish_df[[
            "Ticker", "Close", "Daily_Change_%", "Rel_Vol",
            "Daily_Money_Flow_Score", "Intraday_Net_Flow",
            "Net_Option_Flow", "Composite_Flow_Score"
        ]].to_string(index=False))

    print("\n=== TOP BEARISH OUTFLOW ===")
    if bearish_df.empty:
        print("None")
    else:
        print(bearish_df[[
            "Ticker", "Close", "Daily_Change_%", "Rel_Vol",
            "Daily_Money_Flow_Score", "Intraday_Net_Flow",
            "Net_Option_Flow", "Composite_Flow_Score"
        ]].to_string(index=False))

    print(f"\nSaved Excel report: {Path(OUTPUT_FILE).resolve()}")


if __name__ == "__main__":
    main()